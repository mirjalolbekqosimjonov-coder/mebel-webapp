import json
import logging
from io import BytesIO
from urllib.parse import urlencode

from telegram import (
    Update, InlineKeyboardButton, InlineKeyboardMarkup, WebAppInfo,
)
from telegram.ext import (
    Application, CallbackQueryHandler, CommandHandler,
    ConversationHandler, ContextTypes, MessageHandler, filters,
)

from config import BOT_TOKEN, ADMIN_IDS, WEBAPP_URL
from database import (
    init_db, get_settings, update_setting, save_calc,
    get_recent, get_all, SETTING_LABELS, DEFAULT_SETTINGS,
    get_extra_admins, get_extra_admin_ids, add_admin, remove_admin,
)

logging.basicConfig(format="%(asctime)s - %(levelname)s - %(message)s", level=logging.INFO)
log = logging.getLogger(__name__)

SETTING_SELECT, SETTING_VALUE, ADMIN_ADD = range(3)


# ── Yordamchi ─────────────────────────────────────────────────────────────────
def is_admin(uid: int) -> bool:
    return uid in ADMIN_IDS or uid in get_extra_admin_ids()

def is_super_admin(uid: int) -> bool:
    return uid in ADMIN_IDS

def fmt(n: float) -> str:
    return f"{int(n):,}".replace(",", " ")

def webapp_url() -> str:
    prices = get_settings()
    return f"{WEBAPP_URL}?{urlencode(prices)}"

def admin_kb(uid: int = 0) -> InlineKeyboardMarkup:
    rows = [
        [InlineKeyboardButton("⚙️ Narxlar sozlamasi", callback_data="adm_settings")],
        [
            InlineKeyboardButton("📋 Oxirgi hisoblar", callback_data="adm_history"),
            InlineKeyboardButton("📥 Excel",           callback_data="adm_export"),
        ],
    ]
    if is_super_admin(uid):
        rows.append([InlineKeyboardButton("👥 Adminlar", callback_data="adm_admins")])
    return InlineKeyboardMarkup(rows)

def settings_kb() -> InlineKeyboardMarkup:
    s = get_settings()
    rows = []
    for key in DEFAULT_SETTINGS:
        label = SETTING_LABELS.get(key, key)
        rows.append([InlineKeyboardButton(
            f"{label}: {fmt(s.get(key, 0))} so'm",
            callback_data=f"set_{key}",
        )])
    rows.append([InlineKeyboardButton("🔙 Orqaga", callback_data="set_back")])
    return InlineKeyboardMarkup(rows)


# ── /start ────────────────────────────────────────────────────────────────────
async def cmd_start(update: Update, ctx: ContextTypes.DEFAULT_TYPE) -> None:
    uid = update.effective_user.id
    url = webapp_url()

    kb_rows = [[InlineKeyboardButton("🧮 Hisob-kitob qilish",
                                     web_app=WebAppInfo(url=url))]]
    if is_admin(uid):
        kb_rows.append([InlineKeyboardButton("🔧 Admin panel", callback_data="adm_open")])

    await update.message.reply_text(
        "🪑 *Mebel hisob-kitob botiga xush kelibsiz!*\n\n"
        "Tugmani bosib formani to'ldiring — bot materiallar va foydani hisoblaydi.",
        parse_mode="Markdown",
        reply_markup=InlineKeyboardMarkup(kb_rows),
    )


# ── WebApp dan natija qabul qilish ────────────────────────────────────────────
async def handle_webapp(update: Update, ctx: ContextTypes.DEFAULT_TYPE) -> None:
    try:
        d = json.loads(update.message.web_app_data.data)
    except Exception:
        await update.message.reply_text("❌ Ma'lumotni o'qishda xato.")
        return

    save_calc(
        update.effective_user.id,
        update.effective_user.full_name or update.effective_user.username,
        d.get("furniture"), d.get("material"),
        d.get("width"), d.get("height"), d.get("depth"),
        d.get("extra", ""),
        d.get("body_area", 0), d.get("xdf_area", 0), d.get("edge_m", 0),
        d.get("hinges", 0), d.get("handles", 0), d.get("screw_sets", 1),
        d.get("mat_cost", 0), d.get("sell_price", 0),
        d.get("profit", 0), d.get("profit_pct", 0),
    )

    p_icon = "📈" if d.get("profit", 0) >= 0 else "📉"
    text = (
        f"✅ *Saqlandi!*\n\n"
        f"🪑 {d.get('furniture')} | {d.get('material')}\n"
        f"💰 Material: *{fmt(d.get('mat_cost',0))} so'm*\n"
        f"💵 Sotish: *{fmt(d.get('sell_price',0))} so'm*\n"
        f"{p_icon} Foyda: *{fmt(d.get('profit',0))} so'm* ({d.get('profit_pct',0):.1f}%)"
    )
    await update.message.reply_text(text, parse_mode="Markdown")

    # Adminga xabar
    if ADMIN_IDS:
        admin_msg = (
            f"🔔 Yangi hisob\n"
            f"👤 {update.effective_user.full_name}\n"
            f"🪑 {d.get('furniture')} ({d.get('material')})\n"
            f"💰 {fmt(d.get('mat_cost',0))} → 💵 {fmt(d.get('sell_price',0))}\n"
            f"{p_icon} {fmt(d.get('profit',0))} so'm"
        )
        for aid in ADMIN_IDS:
            try:
                await ctx.bot.send_message(aid, admin_msg)
            except Exception:
                pass


# ── Admin panel ───────────────────────────────────────────────────────────────
async def adm_open(update: Update, ctx: ContextTypes.DEFAULT_TYPE) -> None:
    q = update.callback_query
    await q.answer()
    uid = q.from_user.id
    if not is_admin(uid):
        return
    await q.edit_message_text(
        "🔧 *Admin panel*",
        parse_mode="Markdown",
        reply_markup=admin_kb(uid),
    )


async def adm_callback(update: Update, ctx: ContextTypes.DEFAULT_TYPE) -> int | None:
    q = update.callback_query
    await q.answer()
    if not is_admin(q.from_user.id):
        return None

    if q.data == "adm_settings":
        await q.edit_message_text(
            "⚙️ *Narxlar sozlamasi*\n\nO'zgartirish uchun bosing:",
            parse_mode="Markdown",
            reply_markup=settings_kb(),
        )
        return SETTING_SELECT

    if q.data == "adm_admins":
        await _show_admins(q)
        return None

    if q.data == "adm_history":
        rows = get_recent(15)
        if not rows:
            await q.edit_message_text(
                "📭 Hali hisoblar yo'q.",
                reply_markup=InlineKeyboardMarkup([[
                    InlineKeyboardButton("🔙 Orqaga", callback_data="adm_open")
                ]]),
            )
            return None

        lines = ["📋 *Oxirgi hisoblar:*\n"]
        for row in rows:
            rid, ts, _, uname, fur, mat = row[0], row[1], row[2], row[3], row[4], row[5]
            mat_cost, sell, profit, pct = row[16], row[17], row[18], row[19]
            icon = "📈" if profit >= 0 else "📉"
            lines.append(
                f"🔹 *#{rid}* {ts}\n"
                f"   {fur} ({mat}) | 👤{uname or '—'}\n"
                f"   💰{fmt(mat_cost)} → {icon}{fmt(profit)} ({pct:.0f}%)\n"
            )
        await q.edit_message_text(
            "\n".join(lines),
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup([[
                InlineKeyboardButton("🔙 Orqaga", callback_data="adm_open")
            ]]),
        )
        return None

    if q.data == "adm_export":
        await _send_excel(q)
        return None

    return None


# ── Sozlamalar (ConversationHandler) ─────────────────────────────────────────
async def setting_select(update: Update, ctx: ContextTypes.DEFAULT_TYPE) -> int:
    q = update.callback_query
    await q.answer()

    if q.data == "set_back":
        await q.edit_message_text(
            "🔧 *Admin panel*",
            parse_mode="Markdown",
            reply_markup=admin_kb(),
        )
        return ConversationHandler.END

    key = q.data[4:]
    ctx.user_data["editing_key"] = key
    s = get_settings()
    label = SETTING_LABELS.get(key, key)

    await q.edit_message_text(
        f"✏️ *{label}*\n\nHozirgi: *{fmt(s.get(key, 0))} so'm*\n\nYangi qiymatni kiriting:",
        parse_mode="Markdown",
    )
    return SETTING_VALUE


async def setting_value(update: Update, ctx: ContextTypes.DEFAULT_TYPE) -> int:
    try:
        val = float(update.message.text.replace(",", ".").replace(" ", ""))
        assert val >= 0
        key = ctx.user_data["editing_key"]
        update_setting(key, val)
        label = SETTING_LABELS.get(key, key)
        await update.message.reply_text(
            f"✅ *{label}* yangilandi: *{fmt(val)} so'm*",
            parse_mode="Markdown",
            reply_markup=settings_kb(),
        )
        return SETTING_SELECT
    except Exception:
        await update.message.reply_text("❌ To'g'ri son kiriting (masalan: 45000):")
        return SETTING_VALUE


async def cancel(update: Update, ctx: ContextTypes.DEFAULT_TYPE) -> int:
    await update.message.reply_text("Bekor qilindi.")
    return ConversationHandler.END


# ── Admin boshqaruvi ──────────────────────────────────────────────────────────
def admins_kb() -> InlineKeyboardMarkup:
    rows = []
    for uid, uname, added_at in get_extra_admins():
        label = f"👤 {uname or uid}  ✖ O'chirish"
        rows.append([InlineKeyboardButton(label, callback_data=f"adm_rm_{uid}")])
    rows.append([InlineKeyboardButton("➕ Admin qo'shish", callback_data="adm_add_admin")])
    rows.append([InlineKeyboardButton("🔙 Orqaga",         callback_data="adm_open")])
    return InlineKeyboardMarkup(rows)


async def _show_admins(q) -> None:
    extra = get_extra_admins()
    lines = ["👥 *Adminlar ro'yxati*\n"]
    lines.append("🔑 *Asosiy adminlar (o'zgartirib bo'lmaydi):*")
    for aid in ADMIN_IDS:
        lines.append(f"  • `{aid}`")
    if extra:
        lines.append("\n👤 *Qo'shilgan adminlar:*")
        for uid, uname, added_at in extra:
            lines.append(f"  • {uname or '—'} (`{uid}`) — {added_at}")
    else:
        lines.append("\n_Hali qo'shilgan admin yo'q._")
    await q.edit_message_text(
        "\n".join(lines),
        parse_mode="Markdown",
        reply_markup=admins_kb(),
    )


async def admin_add_start(update: Update, ctx: ContextTypes.DEFAULT_TYPE) -> int:
    q = update.callback_query
    await q.answer()
    if not is_super_admin(q.from_user.id):
        return ConversationHandler.END
    await q.edit_message_text(
        "➕ *Yangi admin qo'shish*\n\n"
        "Foydalanuvchining Telegram ID sini yuboring.\n"
        "_(IDni bilish uchun @userinfobot ga /start yuboring)_",
        parse_mode="Markdown",
        reply_markup=InlineKeyboardMarkup([[
            InlineKeyboardButton("❌ Bekor qilish", callback_data="adm_admins")
        ]]),
    )
    return ADMIN_ADD


async def admin_add_handle(update: Update, ctx: ContextTypes.DEFAULT_TYPE) -> int:
    if not is_super_admin(update.effective_user.id):
        return ConversationHandler.END
    text = update.message.text.strip()
    try:
        new_id = int(text)
    except ValueError:
        await update.message.reply_text("❌ Faqat raqam kiriting (masalan: 123456789):")
        return ADMIN_ADD

    if new_id in ADMIN_IDS:
        await update.message.reply_text("⚠️ Bu foydalanuvchi allaqachon asosiy admin.")
        return ConversationHandler.END

    add_admin(new_id, str(new_id))
    await update.message.reply_text(
        f"✅ `{new_id}` admin sifatida qo'shildi.\n"
        "Endi u admin paneldan foydalana oladi.",
        parse_mode="Markdown",
    )
    return ConversationHandler.END


async def admin_remove(update: Update, ctx: ContextTypes.DEFAULT_TYPE) -> None:
    q = update.callback_query
    await q.answer()
    if not is_super_admin(q.from_user.id):
        return
    uid = int(q.data.split("_")[2])
    remove_admin(uid)
    await _show_admins(q)


# ── Excel eksport ─────────────────────────────────────────────────────────────
async def _send_excel(q) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    rows = get_all()
    if not rows:
        await q.answer("📭 Ma'lumot yo'q!", show_alert=True)
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Hisoblar"
    headers = [
        "ID", "Sana", "Foydalanuvchi", "Mebel", "Material",
        "Kenglik", "Balandlik", "Chuqurlik", "Qo'shimcha",
        "Panel m²", "XDF m²", "Qirra m", "Petlya", "Tutqich", "Vintlar",
        "Material xarajati", "Sotish narxi", "Foyda", "Foyda %",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font      = Font(bold=True, color="FFFFFF")
        cell.fill      = PatternFill("solid", fgColor="1a5276")
        cell.alignment = Alignment(horizontal="center")
    for row in rows:
        ws.append(list(row))
    for col in ws.columns:
        w = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(w + 3, 30)

    out = BytesIO()
    wb.save(out)
    out.seek(0)

    try:
        from telegram import BufferedInputFile
        f = BufferedInputFile(out.read(), filename="mebel_hisoblar.xlsx")
    except ImportError:
        from telegram import InputFile
        f = InputFile(out, filename="mebel_hisoblar.xlsx")

    await q.message.reply_document(f, caption="📊 Excel hisobot tayyor!")


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    init_db()
    app = Application.builder().token(BOT_TOKEN).build()

    settings_conv = ConversationHandler(
        entry_points=[
            CallbackQueryHandler(setting_select,  pattern="^set_"),
            CallbackQueryHandler(admin_add_start, pattern="^adm_add_admin$"),
        ],
        states={
            SETTING_SELECT: [CallbackQueryHandler(setting_select, pattern="^set_")],
            SETTING_VALUE:  [MessageHandler(filters.TEXT & ~filters.COMMAND, setting_value)],
            ADMIN_ADD:      [MessageHandler(filters.TEXT & ~filters.COMMAND, admin_add_handle)],
        },
        fallbacks=[CommandHandler("cancel", cancel)],
        allow_reentry=True,
    )

    app.add_handler(CommandHandler("start", cmd_start))
    app.add_handler(MessageHandler(filters.StatusUpdate.WEB_APP_DATA, handle_webapp))
    app.add_handler(CallbackQueryHandler(adm_open,      pattern="^adm_open$"))
    app.add_handler(CallbackQueryHandler(admin_remove,  pattern="^adm_rm_"))
    app.add_handler(CallbackQueryHandler(adm_callback,  pattern="^adm_"))
    app.add_handler(settings_conv)

    log.info("Bot ishga tushdi...")
    app.run_polling(allowed_updates=Update.ALL_TYPES)


if __name__ == "__main__":
    main()
